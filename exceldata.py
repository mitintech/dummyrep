# import openpyxl module
import openpyxl
# Give the location of the file
path = "C:\\aditya\\demo excel\\hello.xlsx"
# to open the workbook
# workbook object is created
wb = openpyxl.load_workbook(path)
# to select active sheet
sh=wb.active
# select sheet with sheet name
sh = wb["hello"]

# print the total number of rows
print(sh.max_row)
print(sh.max_column)
rw=sh.max_row
cl=sh.max_column
# for i in range(1,rw+1):
#     for j in range(1,cl+1):
#         print(sh.cell(row=i,column=j).value,end="-|-")
#     print()
#
#
# # sh.cell(row=1,column=1).value="MIT Soft"
#
# for i in range(1,rw+1):
#     for j in range(1,cl+1):
#         sh.cell(row=i,column=j).value="Shivani"
#     print()
# wb.save(path)